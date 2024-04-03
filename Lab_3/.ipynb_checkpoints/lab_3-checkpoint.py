import numpy as np
import matplotlib.pyplot as plt
import openpyxl

wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active

x_train_0 = np.array([float(ws.cell(row=i,column=1).value) for i in range(2, ws.max_row+1)])
x_train_1 = np.array([float(ws.cell(row=i,column=2).value) for i in range(2, ws.max_row+1)])
y_train = np.array([float(ws.cell(row=i,column=3).value) for i in range(2, ws.max_row+1)])

# Об'єднуємо дані у вхідні вектори
x_train = np.column_stack((x_train_0, x_train_1, np.ones_like(x_train_0)))