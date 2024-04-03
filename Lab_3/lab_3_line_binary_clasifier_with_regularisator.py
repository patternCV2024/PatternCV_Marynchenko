import numpy as np
import matplotlib.pyplot as plt
import openpyxl

# Сигмоїдна функція втрат
def loss(w, x, y):
    M = np.dot(w, x) * y
    return 2 / (1 + np.exp(M))

# Похідна від сигмоїдальної функції втрат по вектору w
def df(w, x, y):
    L1 = 1.0 # Коефіцієнт L1-регуляризатора
    M = np.dot(w, x) * y
    return -2 * (1 + np.exp(M)) ** (-2) * np.exp(M) * x * y + L1 * np.sign(w)

# Навчальна вибірка з трьома ознаками (третій - константа +1)

wb = openpyxl.load_workbook('lab_3_variant.xlsx')
ws = wb.active

x_train_0 = np.array([float(ws.cell(row=i,column=1).value) for i in range(2, ws.max_row+1)])
x_train_1 = np.array([float(ws.cell(row=i,column=2).value) for i in range(2, ws.max_row+1)])
y_train = np.array([float(ws.cell(row=i,column=3).value) for i in range(2, ws.max_row+1)])

# Об'єднуємо дані у вхідні вектори
x_train = np.column_stack((x_train_0, x_train_1, np.ones_like(x_train_0)))
x_train = [x + [10*x[0], 10*x[1], 5*(x[0]+x[1])] for x in x_train]
x_train = np.array(x_train)

fn = len(x_train[0])
n_train = len(x_train)  # Розмір навчальної вибірки
w = np.zeros(fn)        # Початкові вагові коефіцієнти
nt = 0.00001             # Крок збіжності SGD
lm = 0.01               # Швидкість "забування" для Q
N = 5000                 # Кількість ітерацій SGD

Q = np.mean([loss(x, w, y) for x, y in zip(x_train, y_train)])  # Показник якості
Q_plot = [Q]

# Стохастичний алгоритм градієнтного спуску
for i in range(N):
    k = np.random.randint(0, n_train - 1)       # Випадковий індекс
    ek = loss(w, x_train[k], y_train[k])        # Визначення втрат для обраного вектора
    w = w - nt * df(w, x_train[k], y_train[k])  # Коригування вагів за допомогою SGD
    Q = lm * ek + (1 - lm) * Q                  # Перерахунок показника якості
    Q_plot.append(Q)

Q = np.mean([loss(x, w, y) for x, y in zip(x_train, y_train)]) # Справжнє значення емпіричного ризику після навчання
print(w)
print(Q)

plt.plot(Q_plot)
plt.grid(True)
plt.show()