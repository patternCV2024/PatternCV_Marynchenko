import numpy as np
import matplotlib.pyplot as plt
import openpyxl

wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active

# Читаємо дані з файлу Excel та конвертуємо їх у числовий формат
x_train_0 = np.array([float(ws.cell(row=i,column=1).value) for i in range(2, ws.max_row+1)])
x_train_1 = np.array([float(ws.cell(row=i,column=2).value) for i in range(2, ws.max_row+1)])
y_train = np.array([float(ws.cell(row=i,column=3).value) for i in range(2, ws.max_row+1)])

# Об'єднуємо дані у вхідні вектори
x_train = np.column_stack((x_train_0, x_train_1, np.ones_like(x_train_0)))

# Обчислення вагових коефіцієнтів
w = np.linalg.lstsq(x_train, y_train, rcond=None)[0]
print("Weights:", w)

# Формуємо координати для лінії розділення
line_x = np.arange(0, np.max(x_train[:, 0]) + 1)
line_y = -(w[0] * line_x + w[2]) / w[1]

# Розділяємо дані на класи
x_0 = x_train[y_train == 1]
x_1 = x_train[y_train == -1]

# Відображаємо графік
plt.scatter(x_0[:, 0], x_0[:, 1], color='red')
plt.scatter(x_1[:, 0], x_1[:, 1], color='blue')
plt.plot(line_x, line_y, color='green')
plt.xlim([0, 60])
plt.ylim([0, 51])
plt.ylabel("довжина")
plt.xlabel("ширина")
plt.grid(True)
plt.legend()
plt.show()
